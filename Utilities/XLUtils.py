import openpyxl


class XLUtils:
    @staticmethod
    def get_data_from_excel(file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        data = []

        for i in range(2, sheet.max_row + 1):
            username = sheet.cell(row=i, column=1).value
            password = sheet.cell(row=i, column=2).value
            expected = sheet.cell(row=i, column=3).value
            data.append((username, password, expected))

        return data

    @staticmethod
    def get_row_count(file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def get_column_count(file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_column

    @staticmethod
    def read_data(file, sheet_name, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.cell(row=rownum, column=columnno).value

    @staticmethod
    def write_data(file, sheet_name, rownum, columnno, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        sheet.cell(row=rownum, column=columnno).value = data
        workbook.save(file)
